#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按 JSON 生成 AI 视频分镜表 Excel（分镜表 + 形象设计两个子表）。

用法：
    python scripts/build_storyboard.py --input storyboard.json --output 分镜表.xlsx

说明：
    只负责排版：提示词、画面内容、旁白、音效等内容由调用方（模型）生成后填入 JSON。
    「首帧画面」「尾帧画面」「形象图」列一律留空，由用户在外部 AI 工具生成后填写。
"""

import argparse
import json
import math
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# 分镜表列定义（表头, 列宽, 是否需要文字换行）
STORY_COLS = [
    ("镜号", 6),
    ("时长(秒)", 6),
    ("画面内容", 28),
    ("首帧画面提示词", 40),
    ("首帧画面", 20),
    ("尾帧画面提示词", 40),
    ("尾帧画面", 20),
    ("首尾帧动态提示词", 40),
    ("旁白", 22),
    ("音效", 24),
]

# 形象设计列定义
ASSET_COLS = [
    ("序号", 6),
    ("名称", 28),
    ("提示词", 60),
    ("形象图", 20),
]


def est_lines(text, width):
    """估算一段文字在给定列宽（近似字符单位）下折行后的行数。

    Python 字符串长度按 1 个字符计；中文字符在 Excel 中约占 2 个宽度单位，
    因此每行可容纳的字符数 approx = width / 2。
    """
    if not text:
        return 1
    capacity = max(width / 2.0, 1)
    lines = 0
    for seg in str(text).split("\n"):
        lines += max(1, math.ceil(len(seg) / capacity))
    return lines


def style_row_height(ws, row, text_by_col):
    """依据各文字列折行需求设置行高。"""
    max_lines = 1
    for col_letter, text in text_by_col.items():
        width = ws.column_dimensions[col_letter].width or 10
        max_lines = max(max_lines, est_lines(text, width))
    ws.row_dimensions[row].height = max(20, max_lines * 14 + 6)


def build_storyboard(data):
    wb = Workbook()

    # ---- 分镜表 ----
    ws = wb.active
    ws.title = "分镜表"
    header_fill = PatternFill("solid", fgColor="2F5597")
    single_fill = PatternFill("solid", fgColor="F2F2F2")  # 单图生成：尾帧列浅灰标记
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 表头
    for c, (title, width) in enumerate(STORY_COLS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    shots = data.get("shots", [])
    merge_spans = []  # (start_row, span) 用于合并旁白
    for r, shot in enumerate(shots, start=2):
        index = int(shot.get("index", r - 1))
        duration = shot.get("duration", "")
        scene_content = shot.get("scene_content", "")
        first_prompt = shot.get("first_frame_prompt", "")
        last_prompt = shot.get("last_frame_prompt", "")
        motion_prompt = shot.get("motion_prompt", "")
        narration = shot.get("narration", "")
        sfx = shot.get("sfx", "")
        mode = shot.get("mode", "first_last")
        is_single = (mode == "single") or bool(shot.get("single_frame"))
        if is_single and not last_prompt:
            last_prompt = "同首帧（单图生成）"

        values = [
            f"{index:02d}" if isinstance(index, int) else str(index),
            duration,
            scene_content,
            first_prompt,
            "",  # 首帧画面
            last_prompt,
            "",  # 尾帧画面
            motion_prompt,
            narration,
            sfx,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center_align if c in (1, 2) else body_align
            if is_single and c in (6, 7):
                cell.fill = single_fill

        text_by_col = {
            get_column_letter(3): scene_content,
            get_column_letter(4): first_prompt,
            get_column_letter(6): last_prompt,
            get_column_letter(8): motion_prompt,
            get_column_letter(9): narration,
            get_column_letter(10): sfx,
        }
        style_row_height(ws, r, text_by_col)

        span = int(shot.get("narration_span", 1) or 1)
        if span > 1 and narration:
            merge_spans.append((r, span))

    # 旁白合并
    for start_row, span in merge_spans:
        end_row = start_row + span - 1
        if end_row <= ws.max_row:
            ws.merge_cells(
                start_row=start_row, start_column=9,
                end_row=end_row, end_column=9,
            )
            # 合并后保留首行旁白
            top = ws.cell(row=start_row, column=9)
            top.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- 形象设计 ----
    wa = wb.create_sheet("形象设计")
    for c, (title, width) in enumerate(ASSET_COLS, start=1):
        cell = wa.cell(row=1, column=c, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        wa.column_dimensions[get_column_letter(c)].width = width
    wa.freeze_panes = "A2"
    wa.row_dimensions[1].height = 26

    for r, asset in enumerate(data.get("assets", []), start=2):
        index = asset.get("index", r - 1)
        name = asset.get("name", "")
        prompt = asset.get("prompt", "")
        values = [index, name, prompt, ""]  # 形象图留空
        for c, v in enumerate(values, start=1):
            cell = wa.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = center_align if c == 1 else body_align
        style_row_height(wa, r, {get_column_letter(3): prompt})

    # 文档属性
    if data.get("project_name"):
        wb.properties.title = data["project_name"]
    if data.get("style_name"):
        wb.properties.subject = data["style_name"]

    return wb


def main():
    parser = argparse.ArgumentParser(description="生成 AI 视频分镜表 Excel")
    parser.add_argument("--input", required=True, help="输入 JSON 路径")
    parser.add_argument("--output", required=True, help="输出 .xlsx 路径")
    args = parser.parse_args()

    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = build_storyboard(data)
    wb.save(args.output)
    print(f"已生成：{args.output}")


if __name__ == "__main__":
    main()
